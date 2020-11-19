from openpyxl import Workbook
from typing import List, Dict

_instance = None

class ExcelMgr(object):

    def __init__(self):

        super().__init__()

        self.row = 1

        self.workbook:Workbook = Workbook()

    def saveRow(self, title:str, values:List[str]):

        sheet = self.workbook.active

        sheet.cell(row = self.row, column = 1).value = title

        for colume in range(0, len(values)):
            
            sheet.cell(row = self.row, column = colume + 2).value = values[colume]

        self.row += 1

    def save(self, name:str):

        self.workbook.save(filename=name)


# excelMgr = ExcelMgr()

# excelMgr.saveRow(title="你好", values=["1", "D", "zz"])

# excelMgr.save()